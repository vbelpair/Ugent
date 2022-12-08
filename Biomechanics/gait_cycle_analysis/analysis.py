import openpyxl
import numpy as np
import matplotlib.pyplot as plt

def quit() -> None:
    """_summary_
    Asking the user to prolong the programm.
    """

    q = input("Quiting programm? (y/n): ")
    if q == 'y':
        exit()

def choose(options: list, label: str="options") -> str:
    """_summary_

    Args:
        options (list): List of options.
        label (str, optional): Name of the options. Defaults to "options".

    Returns:
        str: returns choice
    """

    choice = None
    while choice not in options:
        print(f'Choose one of the following {label}: ' + ", ".join(options))
        choice = input('Choice: ')
        if choice not in options:
            print('Invalid option.\n Choose again.')

    return choice

def iterate_ND_range(sheet, start: str, end: str) -> list:
        cells = sheet[start:end]
        if len(cells) == 1:
            return [c1.value for c1 in cells[0]]
        else:
            return [c1[0].value for c1 in cells]
    
def read_in_data_week1(sheet):

    rmax = sheet.max_row
    frames = iterate_ND_range(sheet, 'A6', f'A{rmax}')
    labels = [e for e in iterate_ND_range(sheet, start='A3', end='T3') if e is not None]
    data = {}

    for i in range(0, len(labels)):
        XYZ = np.zeros((rmax-5, 3))
        n = 0
        for row in sheet.iter_rows(min_row=6, min_col=3+3*i, max_row=rmax, max_col=5+3*i):
            m = 0
            for cell in row:
                XYZ[n,m] = cell.value
                m+=1
            n+=1
        data[labels[i]] = XYZ

    return frames, data, labels, 


class GaitCycleAnalysis:

    def __init__(self) -> None:
        while True:
            file = input('Please enter a file for analysis: ')
            print("Opening file...")
            try: 
                self.wb = openpyxl.load_workbook(file)
                break
            except:
                print("File could not be opened. Check wheter the file name/path is correct.")
                print("Try again.")
        print("File has been opened.")
        print("Activating sheet...")
        self.sheetnames = self.wb.sheetnames
        self.sheet = self.wb.active
        self.frames, self.data, self.labels = read_in_data_week1(self.sheet)
        print(f"Sheet {self.sheet} has been opened.")
        
    def app(self):

        while True:
            print(">>> MENU")
            print("0 -> open another sheet")
            print("1 -> calculate statistics")
            print("2 -> plot data")
            print("-1 -> quit")
            options = ["0", "1", "2", "-1"]
            choice = choose(options, label="numbers")

            if choice == "0":
                self.another_sheet()
            elif choice == "1":
                self.calculating_statistics()
            elif choice == "2":
                self.plot_data()
            elif choice == "-1":
                print("Quitting the programm...")
                exit()

    def another_sheet(self):
        print(">>> Opening another sheet.")
        sheetname = choose(self.sheetnames, label="sheetnames")
        print("Activating sheet...")
        self.sheet = self.wb[sheetname]
        self.frames, self.data, self.labels = read_in_data_week1(self.sheet)
        print(f"Sheet {self.sheet} has been opened.")

    def calculating_statistics(self):
        print("Nothing to see here.")

    def plot_data(self):
        print(">>> Plotting the data.")
        print("Choose the data to plot:")
        choice = choose(self.labels)
        values = self.data[choice]
        N, M = values.shape
        ylabels = {0:"x", 1:"y", 2:"z"}
        fig, axes = plt.subplots(nrows=3, figsize=(10,20))
        for m in range(M):
            ax = axes[m]
            ax.plot(self.frames, values[:,m])
            ax.set_xlabel("Frames")
            ax.set_ylabel(f"{ylabels[m]}-track [mm]")
        plt.show()


if __name__ == "__main__":

    print("Welcome to the gait cycle analysis programm!\n")
    object = GaitCycleAnalysis()
    object.app()

        

        





