import uniaxial_analysis as ua
import numpy as np
import os
import openpyxl
import pandas as pd

"""
fine articel abbout multiprocessing and threading in python: https://towardsdatascience.com/multithreading-multiprocessing-python-180d0975ab29

TIP of th day: don't use threading while plotting, use multiprocesses instead
"""

dirname = "C:/Users/Tech/OneDrive/Ugent vincent/2022-2023/Biomechanics/Soft-Tissue-Engineering"
axials = ['uniaxial_hourglass_axial_dataA3', 'uniaxial_hourglass_axial_dataA4', 'uniaxial_hourglass_axial_dataB3', 'uniaxial_hourglass_axial_dataB4', 
        'uniaxial_hourglass_axial_dataC3', 'uniaxial_hourglass_axial_dataC4', 'uniaxial_hourglass_axial_dataD2', 'uniaxial_hourglass_axial_dataD3',] 
circs = ['uniaxial_hourglass_circ_dataA1', 'uniaxial_hourglass_circ_dataA2', 'uniaxial_hourglass_circ_dataB1', 'uniaxial_hourglass_circ_dataB2', 
        'uniaxial_hourglass_circ_dataC1', 'uniaxial_hourglass_circ_dataC2', 'uniaxial_hourglass_circ_dataD1', 'uniaxial_hourglass_circ_dataD2']
straights = ['uniaxial_straight_dataA1', 'uniaxial_straight_dataA2', 'uniaxial_straight_dataB1', 'uniaxial_straight_dataB2', 
            'uniaxial_straight_dataC1', 'uniaxial_straight_dataC2', 'uniaxial_straight_dataD1', 'uniaxial_straight_dataD2']

clinic_data = { 'A': ['F', '68', (49.3,51.0), (99, 149), 26.0, 68, 2.74, 'BAV'], 
                'B': ['F', '74', (58.5,62.1), (70, 143), 37.5, 126, 3.30, '-'], 
                'C': ['M', '58', (46.6,51.2), (65, 135), 30.7, 77, 2.63, 'BAV'], 
                'D': ['M', '55', (38.9,41.4), (73, 144), 30.6, 84, 2.07, 'BAV']}

wdir = os.path.dirname(__file__)
figdir = os.path.join(wdir, 'figures')
# check if figure folder already exists
if os.path.exists(figdir) is False:
    os.mkdir(figdir)
    print(f"Directory {figdir} has been created.")

def main_PK(filename, dirname=dirname, figdir=figdir):

    patient = filename[-2]

    file = os.path.join(dirname, f'Patient_{patient}/' + filename + '.xlsx')

    if os.path.exists(file) is False:
        print(f'No such file {file}')
        return np.NaN

    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active

    # get max number of rows and max number of columns
    row = sheet_obj.max_row
    #column = sheet_obj.max_column

    # get overal information
    L = sheet_obj['B2'].value
    T = sheet_obj['B3'].value
    W = sheet_obj['B4'].value
    #max_tenslie_extension = sheet_obj['B5'].value
    #max_load = sheet_obj['B6'].value

    # get data
    data = np.array([[t.value, l.value, f.value] for t, l, f in sheet_obj['A10': f'C{row}']])
    
    l = data[:,1]
    f = data[:,2]

    if 'circ' in filename:
        A = T*L
        sample_type = 'circular'
    elif 'axial' in filename:
        A = W*T
        sample_type = 'axial'
    else:
        print('No correct filename.')
        return None
    Pult = ua.Pult_1PK(l, f, A, filename, figdir)

    return sample_type, T, Pult

def main_laplace(cd):
        
    cd = 1

    return None

if __name__ == '__main__':

    DF = pd.DataFrame(['patient', 'sample_type', 'T', 'P_ult'])
    for filename in circs + axials:
        results = main_PK(filename)
        sample_type, T, Pult = results[0], results[1], results[2]
        DF[['patient', 'sample', 'T', 'P_ulta']] = [filename[-2], sample_type, T, Pult]
    DF.sort_values(by=['patient', 'sample_type'])
    print(DF)


